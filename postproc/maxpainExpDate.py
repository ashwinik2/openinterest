## openintmaxpain.py calculates max pain of the stock of
## the given exp dates

import csv
import pandas as pd
from pandas import DataFrame
from datetime import date
from datetime import datetime
import time
import os
import sys
import os.path
import datetime
from os import path
import xlsxwriter
import openpyxl
from datetime import timedelta
import chartapi
import openintmi
import Commonapi
import dataapi
import globalheader
import moneyinvExpDate
csv.field_size_limit(sys.maxsize)

contractType =['CALL','PUT']
ifilepath = './../datacolls/output/'
ifilestocklist = './../datacolls/input/'+'StockList.csv'

ofilepath = './output/'
ofilename = ofilepath+'OImax_pain_ExpDate'

#thses datas are used by chart insertion and data insertion of the money invested of call and put
maxpain_row_header = ["CallMPx100", "PutMPx100"]
expdate_chart_y_axis=['MAX_PAIN_CALL x 100','MAX_PAIN_PUT x 100 ']
expdate_chart_col_loc =['2','141']
expdate_chart_insert_pos =["A64","CV64"]


# Adding call/put max pain data of individual strikeprice of an expiration dates of all trading days to excel sheet
# calls insertLineChart which is in chartapi.py 
#writeMaxPainDataToXl(contracttype,wb_obj,max_pain_dict,Symbol):
        #input
        #contracttype - call or put
        #wb_obj - xl file
        # max_pain - max pain dict data
        #symbol - stock expiration symbol
        #StockOptionExpDate - stock expiration date
        
def writeMaxPainDataToXl(contracttype,wb_obj,max_pain_dict,Symbol,StockOptionExpDate):
        if(globalheader.info == 1):
                globalheader.logging.info('writeMaxPainDataToXl Started')
        trading_days_list = max_pain_dict.keys()
        sheet_obj = 0
        strikeprice_list = []

        for expdatekey in range(0,len(trading_days_list)):
                if((contracttype == Commonapi.Contracttype['CALL'].value) and (expdatekey ==0)):
                        addSheetToXl(wb_obj,Symbol,expdatekey,StockOptionExpDate)
                sheets = wb_obj.sheetnames
                sheet_obj = wb_obj[sheets[0]]

                strikeprice_list = max_pain_dict[trading_days_list[expdatekey]].keys()
                strikeprice_value_list = max_pain_dict[trading_days_list[expdatekey]].values()
                if(contracttype == Commonapi.Contracttype.CALL.value):
                        col_index = (int)(expdate_chart_col_loc[contracttype])
                        cellref=sheet_obj.cell(row=1, column=col_index-1)
                        cellref.value=maxpain_row_header[contracttype]
                        cellref=sheet_obj.cell(row=expdatekey+2, column=1)
                        cellref.value=trading_days_list[expdatekey]

                        col_index = (int)(expdate_chart_col_loc[contracttype])
                        for i in range(len(strikeprice_list)):
                                cellref=sheet_obj.cell(row=1, column=col_index+i)
                                cellref.value=strikeprice_list[i]
                        for item in range(len(strikeprice_list)):
                                cellref=sheet_obj.cell(row = expdatekey+2, column=col_index+item)
                                cellref.value=strikeprice_value_list[item]

                else:
                        col_index = (int)(expdate_chart_col_loc[contracttype])
                        cellref=sheet_obj.cell(row=1, column=col_index-1)
                        cellref.value=maxpain_row_header[contracttype]
                        cellref=sheet_obj.cell(row=expdatekey+2, column=col_index-1)
                        cellref.value=trading_days_list[expdatekey]
                                        
                        for i in range(len(strikeprice_list)):
                                cellref=sheet_obj.cell(row=1, column=col_index+i)
                                cellref.value=strikeprice_list[i]

                        for item in range(len(strikeprice_list)):
                            cellref=sheet_obj.cell(row = expdatekey+2, column=col_index+item)
                            cellref.value=strikeprice_value_list[item]
        y_axis_label =[]
        y_axis_label.append(expdate_chart_y_axis[contracttype])
        chart_insert_pos = []
        chart_insert_pos.append(expdate_chart_insert_pos[contracttype])
        chart_col_loc=[]
        chart_col_loc.append(expdate_chart_col_loc[contracttype])
        chartapi.insertLineChart(sheet_obj,strikeprice_list,Symbol,StockOptionExpDate,y_axis_label,chart_col_loc,chart_insert_pos,1)
        if(globalheader.info == 1):
                globalheader.logging.info('writeMaxPainDataToXl Ended')
                
#Adding sheet of each expiration date to xl workbook
#addSheetToXl(wb_obj,symbol,sheetIndex,finalOptionExpDateList):
        #input
        #wb_obj - xl file
        #sheetIndex - sheet index
        #symbol - stock expiration symbol
        #finalOptionExpDateList - stock expiration date list                        
def addSheetToXl(wb_obj,symbol,sheetIndex,finalOptionExpDateList):
        if(globalheader.info == 1):
                globalheader.logging.info('addSheetToXl Started')
        sheet_name =  str(symbol)+finalOptionExpDateList
        if sheet_name in wb_obj.sheetnames:
                if(globalheader.debug == 1):
                        globalheader.logging.debug('%s %s', 'sheet_name present in wb_obj.sheetnames is', sheet_name,wb_obj.sheetnames)                    
        if not sheet_name in wb_obj.sheetnames:
                if(globalheader.debug == 1):
                        globalheader.logging.debug('%s %s', 'sheet_name not present in wb_obj.sheetnames is', sheet_name,wb_obj.sheetnames)                    
                wb_obj.create_sheet(index = sheetIndex , title = sheet_name)
        if(Commonapi.info == 1):
                globalheader.logging.info('addSheetToXl Ended')
        
def maxpain(Symbol,StockOptionExpDate):
        if(globalheader.info == 1):
                globalheader.logging.info('maxpain Started')
        sheet_count = 0
        finalOptionExpDateList =[]
        call_error,istockCSVfilename =  dataapi.getStockCSVFile(0,symbol,ifilepath)
        if(call_error == globalheader.Success): 
                oxlfile = Commonapi.createoFile(symbol,ofilename)
                if os.path.exists(oxlfile):
                        globalheader.logging.warning('%s %s', 'file exists:', oxlfile)
                        os.remove(oxlfile)
                        globalheader.logging.warning('File Removed!')
            
                if os.path.exists(oxlfile):
                        wb_obj = openpyxl.load_workbook(oxlfile)
                        globalheader.logging.warning('%s %s', 'file exists:', oxlfile)

                else:
                        globalheader.logging.warning('%s %s', 'File does not Exists:', oxlfile)
                        wb_obj = openpyxl.Workbook()
  
                for contracttype in range(len(contractType)):                                    
                        call_error,max_pain_dict = Commonapi.getMaxPain(Symbol,StockOptionExpDate,contracttype)
                        if(call_error == globalheader.Success):
                                if(globalheader.info == 1):
                                    globalheader.logging.info('Commonapi.getMaxPain Success')
                                writeMaxPainDataToXl(contracttype,wb_obj,max_pain_dict,Symbol,StockOptionExpDate)
                        else:
                                globalheader.logging.error('%s %d', 'Commonapi.getMaxPain Failed', call_error)
                                return call_error
                wb_obj.save(oxlfile)
                if(globalheader.info == 1):
                        globalheader.logging.info('maxpain Ended')
                return call_error

        else:
                globalheader.logging.error('%s %d', 'max pain Failed', call_error)
                return call_error
                           
if __name__ == "__main__":
        if(globalheader.info == 1):
                globalheader.logging.info('maxpain main Started')
        symbol = raw_input("Enter symbol :") 
        StockOptionExpDate = raw_input("Enter stock exp date in format 'YYMMDD' :",) 
        call_error = maxpain(symbol,StockOptionExpDate)
        if(call_error == globalheader.Success):
                if(globalheader.info == 1):
                    globalheader.logging.info('%s %d','maxpain Success', call_error)
                moneyinvExpDate.money_inv(symbol,StockOptionExpDate)
        else:
                globalheader.logging.error('%s %d', 'maxpain Failed', call_error)
        if(globalheader.info == 1):
                globalheader.logging.info('maxpain main Ended')


                    
